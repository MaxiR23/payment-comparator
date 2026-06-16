# INFO: genera el reporte de Excel con los resultados de la comparación.
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

logger = logging.getLogger(__name__)

# ── Styles ─────────────────────────────────────────────────────────────────────
# borde fino gris que se aplica a TODAS las celdas del reporte.
THIN_BORDER = Border(
    left=Side("thin", color="CCCCCC"),
    right=Side("thin", color="CCCCCC"),
    top=Side("thin", color="CCCCCC"),
    bottom=Side("thin", color="CCCCCC"),
)

# texto blanco y negrita para los encabezados de columna (ej: "Legajo", "Alumno", etc.)
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
# fondo negro para la fila de encabezados de columna
HEADER_FILL = PatternFill("solid", fgColor="000000")
# texto negro grande y negrita para el título principal (ej: "INFORME DE DIFERENCIAS")
TITLE_FONT = Font(bold=True, name="Arial", size=13, color="000000")
# texto normal sin negrita para las filas de datos comunes
NORMAL_FONT = Font(name="Arial", size=10)
# texto rojo y negrita para la celda de diferencia cuando hay discrepancia entre sistemas
DIFF_FONT = Font(bold=True, name="Arial", size=10, color="C0392B")
# fondo rosa suave para resaltar la celda de diferencia
DIFF_FILL = PatternFill("solid", fgColor="FDECEA")
# texto negrita para la fila de totales al pie de la tabla
TOTAL_FONT = Font(bold=True, name="Arial", size=10)
# fondo verde suave para la fila de totales
TOTAL_FILL = PatternFill("solid", fgColor="E8F5E9")
# texto rojo mediano para el subtítulo con el resumen (total diferencia + cantidad de alumnos)
SUBTITLE_FONT = Font(bold=True, name="Arial", size=11, color="C0392B")

# ws: worksheet

# ── Helpers ────────────────────────────────────────────────────────────────────
def _cell(ws, row, col, value, font=None, fill=None, halign="left", fmt=None):
    # escribe un valor en una celda y le aplica fuente, fondo, borde y alineación de forma unificada.
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    return cell

def _write_header_row(ws, row, headers, widths):
    # escribe la fila de encabezados con estilo y ajusta el ancho de cada columna.
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        _cell(ws, row, col, header, font=HEADER_FONT, fill=HEADER_FILL, halign="center")
        ws.column_dimensions[get_column_letter(col)].width = width

# ── Sheets ─────────────────────────────────────────────────────────────────────
def _write_differences(ws, diffs, aulica):
    # genera la hoja principal con el título, los alumnos que tienen diferencias y la fila de totales.
    total_diff = diffs["Diferencia"].sum()

    ws.merge_cells("A1:G1") # une las celdas de A1 hasta G1 para que el título ocupe todo el ancho de la tabla.
    _cell(ws, 1, 1, "INFORME DE DIFERENCIAS — Cupones vs Control", font=TITLE_FONT, halign="center") # escribe el título principal centrado y con estilo.
    ws.row_dimensions[1].height = 28 # aumenta la altura de la fila del título para que se vea mejor.

    ws.merge_cells("A2:G2") # une las celdas de A2 hasta G2 para que el subtítulo ocupe todo el ancho de la tabla.
    _cell(
        ws, 2, 1,
        f"Diferencia total: ${total_diff:,.2f}  |  Alumnos con diferencia: {len(diffs)}",
        font=SUBTITLE_FONT, halign="center", 
    )

    headers = ["Legajo", "Alumno (Control)", "Total Control ($)", "Total Áulica ($)", "Diferencia ($)", "Cupones en Áulica", "Montos Cupones"]
    widths = [10, 35, 18, 18, 18, 38, 38] # anchos personalizados para cada columna, ajustados al contenido esperado.
    _write_header_row(ws, 4, headers, widths) # escribe la fila de encabezados de la tabla de diferencias en la fila 4.

    row = 5

    # para cada alumno con diferencia, busca sus cupones en FUENTE A (a), arma la fila con
    # legajo, nombre, totales y cupones, y escribe cada celda con su estilo:
    # dinero alineado a la derecha, diferencia en rojo, resto en normal.
    for _, r in diffs.iterrows(): 
        coupons = aulica[aulica["Legajo"] == r["Legajo"]][["Nro_Cupon", "Monto"]]
        coupon_ids = ", ".join(coupons["Nro_Cupon"].tolist())
        coupon_amounts = ", ".join(f"${m:,.2f}" for m in coupons["Monto"])

        values = [
            int(r["Legajo"]),
            r["Alumno"],
            r["Total"],
            r["Monto_Aulica"],
            r["Diferencia"],
            coupon_ids,
            coupon_amounts,
        ]

        for col, val in enumerate(values, 1):
            is_money = col in (3, 4, 5)
            is_diff = col == 5
            _cell(
                ws, row, col, val,
                font=DIFF_FONT if is_diff else NORMAL_FONT,
                fill=DIFF_FILL if is_diff else None,
                halign="right" if is_money else "left",
                fmt="#,##0.00" if is_money else None,
            )
        ws.row_dimensions[row].height = 45
        row += 1

    # totals row
    # escribe la fila de totales al pie de la tabla en tres pasos:
    # 1. Columna 1: texto "TOTAL" con fondo verde.
    # 2. Columnas 2, 6 y 7: vacías pero pintadas de verde para completar la fila.
    # 3. Columnas 3, 4 y 5: suma de Control, FUENTE A y diferencia respectivamente.
    # La columna 5 (Diferencia) se resalta en rojo, las otras dos en negrita normal.
    _cell(ws, row, 1, "TOTAL", font=TOTAL_FONT, fill=TOTAL_FILL)
    for col in (2, 6, 7):
        _cell(ws, row, col, None, fill=TOTAL_FILL)
    for col, val in [(3, diffs["Total"].sum()), (4, diffs["Monto_Aulica"].sum()), (5, total_diff)]:
        _cell(
            ws, row, col, val,
            font=DIFF_FONT if col == 5 else TOTAL_FONT,
            fill=TOTAL_FILL, halign="right", fmt="#,##0.00",
        )

# genera la hoja secundaria con los alumnos que están en Control pero no tienen cupones en FUENTE A.
def _write_control_only(ws, control_only):
    ws.merge_cells("A1:D1") 
    ws.row_dimensions[1].height = 28
    _cell(
        ws, 1, 1,
        "Alumnos presentes en Control pero SIN cupones en Áulica",
        font=Font(bold=True, name="Arial", size=12, color="000000"),
        halign="center",
    )

    headers = ["Legajo", "Alumno", "Cuota Abril ($)", "Total ($)"]
    widths = [10, 42, 18, 18]
    _write_header_row(ws, 3, headers, widths)

    for i, (_, r) in enumerate(control_only.iterrows(), 4):
        values = [
            int(r["Legajo"]) if pd.notna(r["Legajo"]) else "",
            r["Alumno"],
            r["Cuota_Abril"],
            r["Total"],
        ]
        for col, val in enumerate(values, 1):
            is_money = col in (3, 4)
            _cell(
                ws, i, col, val,
                font=NORMAL_FONT,
                halign="right" if is_money else "left",
                fmt="#,##0.00" if is_money else None,
            )

# ── Public API ─────────────────────────────────────────────────────────────────
# punto de entrada del reporte: crea el libro Excel, genera las dos hojas y guarda el archivo en disco.
def write_report(
    diffs: pd.DataFrame,
    control_only: pd.DataFrame,
    aulica: pd.DataFrame,
    output_path: str,
) -> None:
    wb = Workbook()

    ws_diffs = wb.active
    ws_diffs.title = "Diferencias"
    _write_differences(ws_diffs, diffs, aulica)

    ws_ctrl = wb.create_sheet("Solo en Control")
    _write_control_only(ws_ctrl, control_only)

    # genera un nombre único con timestamp: diferencias_260415_181535.xlsx
    timestamp = datetime.now().strftime("%y%m%d_%H%M%S")
    output_path = output_path.replace(".xlsx", f"_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("Report saved: %s", output_path)